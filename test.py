from openpyxl.styles import Font, Alignment
wb = openpyxl.Workbook()

# Xóa sheet mặc định (nếu cần)
if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

# Tạo và đặt tên sheet 1: Users
sheet_users = wb.create_sheet("Users")
# Thêm tiêu đề
sheet_users.append(["ID", "Username", "Email", "Role ID"])
# Định dạng tiêu đề
for cell in sheet_users[1]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")

# Dữ liệu mẫu cho sheet Users
users_data = [
    (1, "director1", "director1@example.com", 1),
    (2, "vice_director1", "vice_director1@example.com", 2),
    (3, "accountant1", "accountant1@example.com", 3)
]
# Thêm dữ liệu vào sheet Users
for user in users_data:
    sheet_users.append(user)

# Tạo và đặt tên sheet 2: Roles
sheet_roles = wb.create_sheet("Roles")
# Thêm tiêu đề
sheet_roles.append(["ID", "Role Name", "Description"])
# Định dạng tiêu đề
for cell in sheet_roles[1]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")

# Dữ liệu mẫu cho sheet Roles
roles_data = [
    (1, "Director", "Full access to all functions"),
    (2, "Vice Director", "View-only access"),
    (3, "Accountant", "View and edit financial data")
]
# Thêm dữ liệu vào sheet Roles
for roles in roles_data:
    sheet_roles.append(roles)

# Lưu file Excel
file_path = "system_data.xlsx"
wb.save(file_path)
print(f"File Excel đã được tạo tại: {file_path}")
```