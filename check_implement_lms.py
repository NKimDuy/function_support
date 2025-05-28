from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.service import Service
import time
import requests
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
import re
import pandas as pd

#------------------------------
# khớp dữ liệu lấy từ lsa
#------------------------------
def get_dictionary(txt):
      lines = txt.splitlines()
      subject = ""
      group = ""
      techer_id = ""
      teacher_name = ""
      for i, line in enumerate(lines):
            if line.startswith("[242]"):
                  match = re.search(r"\[242\]\s+(\w+)\s+-.*\((\w+)-(\w+)\)", line)
                  if match:
                        subject = match.group(1)     # FINA2343
                        techer_id = match.group(2)      # KT196
                        group = match.group(3)       # TN303
            elif line.startswith("Giảng viên"):
                  teacher_name = line.split(". ", 1)[-1].strip()
      return [group, subject, techer_id, teacher_name]


#------------------------------
# lấy dữ liệu từ file excel theo từng học kỳ
# row[0]: nhóm
# row[1]: mã môn học
# row[2]: mã giảng viên
# row[3]: tên giảng viên
# row[4]: tên khoa
#------------------------------
def get_detail(file):
      list_detail = {}
      wb = load_workbook(file)
      ws = wb.active

      for row in ws.iter_rows(values_only=True):
            key = row[0] + "-" + row[1]
            list_detail[key] = [row[2], row[3], row[4]]
      wb.close()
      return list_detail


#------------------------------
# vào trang web lsa và lấy dữ liệu giảng viên đã soạn bài về 
#------------------------------
def get_lsa(semester, url):
      value = "Đăng nhập bằng HCMCOU-SSO"
      semester = " ".join(["[LIVE] LMS TX", semester])
      general = "http://lsa.ou.edu.vn/vi/admin/mm/report/usersiteoverviews"
      xpath = f"//button[text()='{value}']"
      chrome_options = Options()
      chrome_options = Options()
      chrome_options.add_argument("--headless=new")  # Chế độ headless mới
      chrome_options.add_argument("--disable-blink-features=AutomationControlled")  # Tắt automation
      chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36")
      chrome_options.add_argument("--window-size=1920,1080")
      chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
      chrome_options.add_experimental_option("useAutomationExtension", False)

      # Khởi tạo driver
      driver = webdriver.Chrome(options=chrome_options)

      # Che giấu navigator.webdriver
      driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
      "source": """
            Object.defineProperty(navigator, 'webdriver', {
                  get: () => undefined
            });
      """
      })
      driver = webdriver.Chrome(options=chrome_options)
      #driver = webdriver.Chrome()
      #driver.get("http://lsa.ou.edu.vn")
      driver.get(url)

      try:
            button_semester = WebDriverWait(driver, 15).until(
                  EC.element_to_be_clickable((By.XPATH, xpath))
            )
            button_semester.click()
      except:
            print("Không tìm thấy nút đăng nhập bằng HCMCOU-SSO")

      try:
            dropdown = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, "form-usertype"))
            )
            select_type_user = Select(dropdown)
            select_type_user.select_by_visible_text("Cán bộ-Nhân viên / Giảng viên")
      except:
            print("Không tìm thấy nút Cán bộ nhân viên/Giảng viên")

      try:
            username = WebDriverWait(driver, 15).until(
                  EC.presence_of_element_located((By.ID, "form-username"))
            )
            username.send_keys("duy.nk")
      except:
            print("Không tìm thấy ô để nhập tài khoản")

      try:
            password = driver.find_element(By.ID, "form-password")
            password.send_keys("tonyTeo!998")
      except:
            print("Không tìm thấy ô để nhập mật khẩu")

      try:
            captcha = driver.find_element(By.ID, "form-captcha")
            captcha.send_keys("clcl")
      except:
            print("Không tìm thấy ô để nhập Capcha")

      try:
            button_login = WebDriverWait(driver, 15).until(
                  EC.element_to_be_clickable((By.XPATH, "//button[text()='Đăng nhập']"))
            )
            button_login.click()
      except:
            print("Không tìm thấy nút đăng nhập")

      try:
            has_found_button_allow = driver.find_elements(By.CSS_SELECTOR, ".btn.btn-success.btn-approve")
            if has_found_button_allow:
                  button_allow = WebDriverWait(driver, 15).until(
                  EC.presence_of_element_located((By.CSS_SELECTOR, ".btn btn-success btn-approve"))
                  )
                  button_allow.click()
      except:
            print("Không tìm thấy nút để nhấn đồng ý")

      try:
            dropdown_semester = WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.ID, "moodlesiteid"))
            )

            select_type_semester = Select(dropdown_semester)
            select_type_semester.select_by_value("54")
      except:
            print("Không tìm thấy dropdownlist thể hiện học kỳ")

      try:
            driver.execute_script("arguments[0].style.display='block';", driver.find_element(By.ID, "menu_1_sub"))
            overview_link = WebDriverWait(driver, 20).until(
                  EC.element_to_be_clickable((By.CSS_SELECTOR, "a[href*='usersiteoverviews']"))
            )
            overview_link.click()
      except:
            print("Không tìm thấy nút report")

      try:
            table = WebDriverWait(driver, 40).until(
                  EC.presence_of_element_located((By.ID, "ourptlistcourse"))  # Thay "myTable" bằng ID thực tế
            )
      except:
            print("không tìm thấy bảng")

      try:
            rows = WebDriverWait(driver, 20).until(
                  EC.presence_of_all_elements_located((By.XPATH, ".//tr"))
            )
      except: 
            print("không tìm thấy các dòng")

      get_subject = []
      for row in rows:
            cells = row.find_elements(By.XPATH, ".//td")
            for cell in cells:
                  group, subject, teacher_id, teacher_name = get_dictionary(cell.text)
                  if group and subject:
                        get_subject.append(group + "-" + subject)
      return get_subject


#------------------------------
# lấy dữ liệu từ api gồm danh sách môn học bắt đầu từ ngày đến ngày
#------------------------------
def get_subject_by_day(semester, from_day, to_day, file):
      url_link_unit = "https://api.ou.edu.vn/api/v1/hdmdp"
      url_list_subject_semester = "https://api.ou.edu.vn/api/v1/tkblopdp"
      headers = {
            "Authorization": "Bearer 52C4E470AF3AE6C56276FAE8666788291F7AEA1667FE67C9DF743FF49FD5C74B"
      }
      from_day = datetime.strptime(from_day, "%Y-%m-%d")
      to_day = datetime.strptime(to_day, "%Y-%m-%d")
      list_subject_in_range = {}

      get_list_unit = requests.get(url_link_unit, headers=headers)
      list_unit = get_list_unit.json()
      for unit in list_unit.get("data", []):
            params_list_subject_semester = {
                  "nhhk": semester,
                  "madp": unit["MaDP"]
            }
            get_list_subject_semester = requests.get(url_list_subject_semester, headers=headers, params=params_list_subject_semester)
            list_subject_semester = get_list_subject_semester.json()
            for lst in list_subject_semester.get("data", []):
                  if lst["TUNGAYTKB"] is not None:
                        if from_day <= datetime.strptime(lst["TUNGAYTKB"], "%Y-%m-%d") <= to_day:
                              key = lst["NhomTo"] + "-" + lst["MaMH"]
                              list_detail = get_detail(file)
                              if key not in list_subject_in_range:
                                    list_subject_in_range[key] =  [
                                          lst["NhomTo"], 
                                          lst["MaMH"],
                                          lst["TenMH"],
                                          lst["TUNGAYTKB"],
                                          lst["MaLop"],
                                          lst["TenLop"],
                                          lst["MaDP"],
                                          lst["TenDP"],
                                         list_detail[key][0], #mã giảng viên
                                         list_detail[key][1], #tên giảng viên
                                         list_detail[key][2]  #khoa
                                    ]
                              else:
                                    list_subject_in_range[key][4] = ",".join([list_subject_in_range[key][4], lst["MaLop"]])     
      return list_subject_in_range


#------------------------------
# Tạo file báo cáo
#------------------------------
def create_file_report(data, from_day, to_day, semester):

      wb = Workbook()
      if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
      sheet_general = wb.create_sheet("Tổng quan")
      sheet_detail = wb.create_sheet("Chi tiết")

      # TODO: Định dạng cho các dòng trong file
      header_font = Font(name="Times New Roman", size=12, bold=True)
      footer_font = Font(name="Times New Roman", size=12, bold=True)
      header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
      header_fill = PatternFill(start_color="97FFFF", end_color="97FFFF", fill_type="solid")
      data_font = Font(name="Times New Roman", size=11)
      data_alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
      border_style = Border(
            left = Side(style="thin"),
            right = Side(style="thin"),
            top = Side(style="thin"),
            bottom = Side(style="thin")
      )

      # TODO: Tạo sheet thứ 1 về tổng quản tình hình của từng khoa 
      list_department = []
      list_sum_department = []
      list_has_lms = []
      list_none_lms = []
      # TODO: tạo 3 mảng gồm khoa, tổng số nhóm, có thực hiện LMS, không có LMS
      for row in data:
            if row["department"] not in list_department:
                  list_department.append(row["department"])
                  list_sum_department.append(1)
                  list_has_lms.append(1 if row["has_lms"] != "" else 0)
                  list_none_lms.append(0 if row["has_lms"] != "" else 1)
            else:
                  idx_department = list_department.index(row["department"])
                  list_sum_department[int(idx_department)] += 1
                  if row["has_lms"] != "":
                        list_has_lms[idx_department] += 1
                  else:
                        list_none_lms[idx_department] += 1

      # TODO: Thêm dòng tiêu đề cho sheet tổng quan
      header_general_department = sheet_general.cell(row=1, column=1)
      header_general_department.value = "Khoa"
      header_general_department.font = header_font
      header_general_department.fill = header_fill
      header_general_department.alignment = header_alignment
      header_general_department.border = border_style

      header_general_sum_subject = sheet_general.cell(row=1, column=2)
      header_general_sum_subject.value = "Tổng số nhóm môn học"
      header_general_sum_subject.font = header_font
      header_general_sum_subject.fill = header_fill
      header_general_sum_subject.alignment = header_alignment
      header_general_sum_subject.border = border_style                                  
                                                      
      header_general_has_lms = sheet_general.cell(row=1, column=3)
      header_general_has_lms.value = "Đã soạn LMS"
      header_general_has_lms.font = header_font
      header_general_has_lms.fill = header_fill
      header_general_has_lms.alignment = header_alignment
      header_general_has_lms.border = border_style           

      header_general_none_lms = sheet_general.cell(row=1, column=4)
      header_general_none_lms.value = "Chưa soạn"
      header_general_none_lms.font = header_font
      header_general_none_lms.fill = header_fill
      header_general_none_lms.alignment = header_alignment
      header_general_none_lms.border = border_style    

      # TODO: Thêm dữ liệu thống kê của từng khoa
      sum_department = 0
      sum_has_lms = 0
      sum_none_lms = 0
      for depart_idx, depart in enumerate(list_department, start=2):
            body_general_department = sheet_general.cell(row=depart_idx, column=1)
            body_general_department.value = depart
            body_general_department.font = data_font
            body_general_department.alignment = data_alignment
            body_general_department.border = border_style

            # TODO depart_idx - 2: vì phải bắt đầu đổ dữ liệu ở dòng số 2, nhưng chỉ mục của list cần lấy là 0, nên cần trừ đi 2 để lấy đúng chỉ mục

            body_general_sum_subject =  sheet_general.cell(row=depart_idx, column=2)
            body_general_sum_subject.value = list_sum_department[depart_idx - 2]
            body_general_sum_subject.font = data_font
            body_general_sum_subject.alignment = data_alignment
            body_general_sum_subject.border = border_style
            sum_department += int(list_sum_department[depart_idx - 2]) # lấy tổng số nhóm môn học của tất cả các khoa

            body_general_has_lms =  sheet_general.cell(row=depart_idx, column=3)
            body_general_has_lms.value = list_has_lms[depart_idx - 2]
            body_general_has_lms.font = data_font
            body_general_has_lms.alignment = data_alignment
            body_general_has_lms.border = border_style
            sum_has_lms += int(list_has_lms[depart_idx - 2]) # lấy tổng số môn học có LMS

            body_general_none_lms = sheet_general.cell(row=depart_idx, column=4)
            body_general_none_lms.value = list_none_lms[depart_idx - 2]
            body_general_none_lms.font = data_font
            body_general_none_lms.alignment = data_alignment
            body_general_none_lms.border = border_style
            sum_none_lms += int(list_none_lms[depart_idx - 2]) # lấy tổng số môn học không có lms
      
      # TODO: thêm dòng tổng kết ở cuối sheet tổng quan
      end_header_general_department = sheet_general.cell(row=len(list_department) + 2, column=1)
      end_header_general_department.value = "Tổng cộng"
      end_header_general_department.font = footer_font
      end_header_general_department.alignment = data_alignment
      end_header_general_department.border = border_style

      end_header_general_sum_department = sheet_general.cell(row=len(list_department) + 2, column=2)
      end_header_general_sum_department.value = sum_department
      end_header_general_sum_department.font = footer_font
      end_header_general_sum_department.alignment = data_alignment
      end_header_general_sum_department.border = border_style

      end_header_general_has_lms = sheet_general.cell(row=len(list_department) + 2, column=3)
      end_header_general_has_lms.value = sum_has_lms
      end_header_general_has_lms.font = footer_font
      end_header_general_has_lms.alignment = data_alignment
      end_header_general_has_lms.border = border_style

      end_header_general_none_lms = sheet_general.cell(row=len(list_department) + 2, column=4)
      end_header_general_none_lms.value = sum_none_lms
      end_header_general_none_lms.font = footer_font
      end_header_general_none_lms.alignment = data_alignment
      end_header_general_none_lms.border = border_style
      
      # TODO: thiếp lập cột khoa về phía bên trái và điều chỉnh độ rộng của cột khoa
      max_length_header = 0
      column_department = get_column_letter(1)
      for cell in sheet_general[column_department][1:]:
            cell.alignment = Alignment(horizontal="general", vertical="center", wrap_text=True)
      for cell in sheet_general[column_department]:
            if len(str(cell.value)) > max_length_header:
                  max_length_header = len(str(cell.value))
            adjust_width_department_general = max_length_header + 5
            sheet_general.column_dimensions[column_department].width = adjust_width_department_general


      # TODO: Tạo sheet thứ 2 về chi tiết từng môn học của từng khoa
      title = [
            "STT",
            "Khoa phụ trách",
            "Mã địa phương",
            "Tên địa phương",
            "Mã môn học",
            "Tên môn học",
            "Mã nhóm",
            "Mã lớp",
            "Tên lớp",
            "Mã giảng viên",
            "Tên giảng viên",
            "Ngày bắt đầu",
            "Đã soạn LMS"
      ]

      # TODO: thiết lập giá trị và định dạng cho dòng tiêu đề
      for title_idx, row_title in enumerate(title, start = 1):
            header_detail = sheet_detail.cell(row = 1, column = title_idx)
            header_detail.value = row_title
            header_detail.font = header_font
            header_detail.alignment = header_alignment
            header_detail.fill = header_fill
            header_detail.border = border_style
      
      # TODO: thiết lập giá trị và định dạng cho các dòng còn lại
      for row_idx, row_data in enumerate(data, start = 2):
            has_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid") # mặc định các dòng trong body sẽ là màu trắng
            if row_data["has_lms"] != "x":
                  has_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # nếu có môn chưa có lms thì sẽ được tô vàng
            
            body_detail_id = sheet_detail.cell(row = row_idx, column = 1)
            body_detail_id.value = row_idx - 1
            body_detail_id.font = data_font
            body_detail_id.alignment = data_alignment
            body_detail_id.border = border_style
            body_detail_id.fill = has_fill

            body_detail_department = sheet_detail.cell(row = row_idx, column = 2)
            body_detail_department.value = row_data["department"]
            body_detail_department.font = data_font
            body_detail_department.alignment = data_alignment
            body_detail_department.border = border_style
            body_detail_department.fill = has_fill

            body_detail_id_unit = sheet_detail.cell(row = row_idx, column = 3)
            body_detail_id_unit.value = row_data["id_unit"]
            body_detail_id_unit.font = data_font
            body_detail_id_unit.alignment = data_alignment
            body_detail_id_unit.border = border_style
            body_detail_id_unit.fill = has_fill

            body_detail_name_unit = sheet_detail.cell(row = row_idx, column = 4)
            body_detail_name_unit.value = row_data["name_unit"]
            body_detail_name_unit.font = data_font
            body_detail_name_unit.alignment = data_alignment
            body_detail_name_unit.border = border_style
            body_detail_name_unit.fill = has_fill

            body_detail_id_subject = sheet_detail.cell(row = row_idx, column = 5)
            body_detail_id_subject.value = row_data["id_subject"]
            body_detail_id_subject.font = data_font
            body_detail_id_subject.alignment = data_alignment
            body_detail_id_subject.border = border_style
            body_detail_id_subject.fill = has_fill

            body_detail_name_subject = sheet_detail.cell(row = row_idx, column = 6)
            body_detail_name_subject.value = row_data["name_subject"]
            body_detail_name_subject.font = data_font
            body_detail_name_subject.alignment = data_alignment
            body_detail_name_subject.border = border_style
            body_detail_name_subject.fill = has_fill

            body_detail_group = sheet_detail.cell(row = row_idx, column = 7)
            body_detail_group.value = row_data["group"]
            body_detail_group.font = data_font
            body_detail_group.alignment = data_alignment
            body_detail_group.border = border_style
            body_detail_group.fill = has_fill

            body_detail_id_class = sheet_detail.cell(row = row_idx, column = 8)
            body_detail_id_class.value = row_data["id_class"]
            body_detail_id_class.font = data_font
            body_detail_id_class.alignment = data_alignment
            body_detail_id_class.border = border_style
            body_detail_id_class.fill = has_fill

            body_detail_name_class = sheet_detail.cell(row = row_idx, column = 9)
            body_detail_name_class.value = row_data["name_class"]
            body_detail_name_class.font = data_font
            body_detail_name_class.alignment = data_alignment
            body_detail_name_class.border = border_style
            body_detail_name_class.fill = has_fill

            body_detail_id_teacher = sheet_detail.cell(row = row_idx, column = 10)
            body_detail_id_teacher.value = row_data["id_teacher"]
            body_detail_id_teacher.font = data_font
            body_detail_id_teacher.alignment = data_alignment
            body_detail_id_teacher.border = border_style
            body_detail_id_teacher.fill = has_fill

            body_detail_name_teacher = sheet_detail.cell(row = row_idx, column = 11)
            body_detail_name_teacher.value = row_data["name_teacher"]
            body_detail_name_teacher.font = data_font
            body_detail_name_teacher.alignment = data_alignment
            body_detail_name_teacher.border = border_style
            body_detail_name_teacher.fill = has_fill

            body_detail_from_day = sheet_detail.cell(row = row_idx, column = 12)
            body_detail_from_day.value = row_data["from_day"]
            body_detail_from_day.font = data_font
            body_detail_from_day.alignment = data_alignment
            body_detail_from_day.border = border_style
            body_detail_from_day.fill = has_fill

            body_detail_has_lms = sheet_detail.cell(row = row_idx, column = 13)
            body_detail_has_lms.value = row_data["has_lms"]
            body_detail_has_lms.font = data_font
            body_detail_has_lms.alignment = data_alignment
            body_detail_has_lms.border = border_style
            body_detail_has_lms.fill = has_fill

      # TODO: điều chỉnh độ rộng của cột dựa trên giá trị dài nhất
      for col_idx in range(1, len(data[0]) + 1):
            max_length = 0
            column = get_column_letter(col_idx)
            for cell in sheet_detail[column]:
                  if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            adjust_width = max_length + 2
            sheet_detail.column_dimensions[column].width = adjust_width
      
      wb.save(f"{semester} - Tình hình soạn thảo LMS từ ngày {datetime.strptime(from_day, "%Y-%m-%d").strftime("%d-%m-%Y")} đến ngày {datetime.strptime(to_day, "%Y-%m-%d").strftime("%d-%m-%Y")}.xlsx")


def main():
      # data = [
      #       {"group": "SG001", "id_subject": "ACCO4331", "name_subject": "Quản trị học", "from_day": "31/12/1994", "id_class": "TM123456", "name_class": "Lớp luật Đồng Tháp Mười", "id_unit": "TM", "name_unit": "Đồng Tháp Mười Long An", "id_teacher": "TX001", "name_teacher": "Nguyễn Kim Duy", "department": "Luật", "has_lms": "x"},
      #       {"group": "SG001", "id_subject": "ACCO4331", "name_subject": "Quản trị học", "from_day": "31/12/1994", "id_class": "TM123456", "name_class": "Lớp luật Đồng Tháp Mười", "id_unit": "TM", "name_unit": "Đồng Tháp Mười Long An", "id_teacher": "TX001", "name_teacher": "Nguyễn Kim Duy", "department": "Luật", "has_lms": "x"},
      #       {"group": "SG001", "id_subject": "ACCO4331", "name_subject": "Quản trị học", "from_day": "31/12/1994", "id_class": "TM123456", "name_class": "Lớp luật Đồng Tháp Mười", "id_unit": "TM", "name_unit": "Đồng Tháp Mười Long An", "id_teacher": "TX001", "name_teacher": "Nguyễn Kim Duy", "department": "Luật", "has_lms": ""},
      #       {"group": "SG001", "id_subject": "ACCO4331", "name_subject": "Quản trị học", "from_day": "31/12/1994", "id_class": "TM123456", "name_class": "Lớp luật Đồng Tháp Mười", "id_unit": "TM", "name_unit": "Đồng Tháp Mười Long An", "id_teacher": "TX001", "name_teacher": "Nguyễn Kim Duy", "department": "Xây dựng", "has_lms": "x"},
      #       {"group": "SG001", "id_subject": "ACCO4331", "name_subject": "Quản trị học", "from_day": "31/12/1994", "id_class": "TM123456", "name_class": "Lớp luật Đồng Tháp Mười", "id_unit": "TM", "name_unit": "Đồng Tháp Mười Long An", "id_teacher": "TX001", "name_teacher": "Nguyễn Kim Duy", "department": "Xây dựng", "has_lms": ""},
      #       {"group": "SG001", "id_subject": "ACCO4331", "name_subject": "Quản trị học", "from_day": "31/12/1994", "id_class": "TM123456", "name_class": "Lớp luật Đồng Tháp Mười", "id_unit": "TM", "name_unit": "Đồng Tháp Mười Long An", "id_teacher": "TX001", "name_teacher": "Nguyễn Kim Duy", "department": "Quản trị kinh doanh", "has_lms": "x"}
      # ]
      #create_file_report(data)
      
      semester = "242"
      from_day = "2025-05-19"
      to_day = "2025-05-25"
      url_lsa = "http://lsa.ou.edu.vn"
      file = "242_detail.xlsx"
      report_final = []
      list_lsa = get_lsa(semester, url_lsa)
      list_subject_by_day = get_subject_by_day(semester, from_day, to_day, file)

      for key, value in list_subject_by_day.items():
            temp = {}
            temp["group"] = value[0]
            temp["id_subject"] = value[1]
            temp["name_subject"] = value[2]
            temp["from_day"] = value[3]
            temp["id_class"] = value[4]
            temp["name_class"] = value[5]
            temp["id_unit"] = value[6]
            temp["name_unit"] = value[7]
            temp["id_teacher"] = value[8]
            temp["name_teacher"] = value[9]
            temp["department"] = value[10]
            temp["has_lms"] = "x" if key in list_lsa else ""
            report_final.append(temp)
      create_file_report(report_final, from_day, to_day, semester)
if __name__ == "__main__":
      main()